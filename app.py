from flask import Flask, request, send_file, jsonify
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from flask_cors import CORS
import requests
import difflib
import os

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Fixed live URLs for latest comparison
FIXED_URLS = {
    "skilled_worker": "https://www.gov.uk/guidance/immigration-rules/immigration-rules-appendix-skilled-worker",
    "sponsor_documents": "https://www.gov.uk/government/publications/supporting-documents-for-sponsor-applications-appendix-a/appendix-a-supporting-documents-for-sponsor-licence-application-accessible-version"
}

# --- Fetch latest archives from Wayback Machine ---
def get_last_archives(original_url, count=3):
    cdx_api = "http://web.archive.org/cdx/search/cdx"
    params = {
        "url": original_url,
        "output": "json",
        "limit": count,
        "filter": "statuscode:200",
        "collapse": "digest",
        "fl": "timestamp,original",
        "sort": "reverse"
    }
    response = requests.get(cdx_api, params=params)
    response.raise_for_status()

    data = response.json()
    if len(data) <= 1:
        return []

    # Skip header
    archives = []
    for entry in data[1:]:
        timestamp, url = entry
        archives.append(f"https://web.archive.org/web/{timestamp}/{url}")

    return archives

# --- Updated GET /get-doc-types using dynamic archive fetch ---
@app.route('/get-doc-types', methods=['GET'])
def get_doc_types():
    try:
        skilled_worker_archives = get_last_archives(FIXED_URLS["skilled_worker"])
        sponsor_documents_archives = get_last_archives(FIXED_URLS["sponsor_documents"])

        doc_types = {
            "skilled_worker": skilled_worker_archives,
            "sponsor_documents": sponsor_documents_archives
        }

        return jsonify(doc_types)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --- Extract plain text from a webpage ---
def get_text_from_url(url):
    response = requests.get(url)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, 'html.parser')
    for tag in soup(['script', 'style']):
        tag.decompose()
    return soup.get_text(separator='\n', strip=True)

# --- Compare two texts line by line ---
def compare_and_align_lines(old_text, new_text):
    old_lines = old_text.splitlines()
    new_lines = new_text.splitlines()
    sm = difflib.SequenceMatcher(None, old_lines, new_lines)
    result = []

    for opcode, i1, i2, j1, j2 in sm.get_opcodes():
        if opcode == 'equal':
            for i in range(i2 - i1):
                result.append((old_lines[i1 + i], new_lines[j1 + i], 'Unchanged'))
        elif opcode == 'replace':
            max_len = max(i2 - i1, j2 - j1)
            for k in range(max_len):
                old_line = old_lines[i1 + k] if k < (i2 - i1) else ''
                new_line = new_lines[j1 + k] if k < (j2 - j1) else ''
                status = 'Changed' if old_line and new_line else 'Removed' if old_line else 'Added'
                result.append((old_line, new_line, status))
        elif opcode == 'delete':
            for i in range(i1, i2):
                result.append((old_lines[i], '', 'Removed'))
        elif opcode == 'insert':
            for j in range(j1, j2):
                result.append(('', new_lines[j], 'Added'))

    return result

# --- Save comparison to Excel ---
def save_to_excel(data, save_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    headers = ['Line No.', 'Archived Text', 'Current Text', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")

    for idx, (old, new, status) in enumerate(data, start=1):
        ws.append([idx, old, new, status])

    wb.save(save_path)

# --- POST /compare: Manual Comparison of Any Two URLs ---
@app.route('/compare', methods=['POST'])
def compare():
    data = request.json
    url1 = data.get('archived_url')
    url2 = data.get('current_url')

    if not url1 or not url2:
        return jsonify({'error': 'Both URLs are required'}), 400

    try:
        text1 = get_text_from_url(url1)
        text2 = get_text_from_url(url2)
        comparison = compare_and_align_lines(text1, text2)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f'comparisons/diff_{timestamp}.xlsx'
        os.makedirs('comparisons', exist_ok=True)
        save_to_excel(comparison, file_path)

        return send_file(file_path, as_attachment=True)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- POST /compare-fixed: Compare archive vs fixed doc URL ---
@app.route('/compare-fixed', methods=['POST'])
def compare_fixed():
    data = request.json
    archived_url = data.get('archived_url')
    doc_type = data.get('doc_type')

    if not archived_url or doc_type not in FIXED_URLS:
        return jsonify({'error': 'Archived URL and valid doc_type are required'}), 400

    try:
        current_url = FIXED_URLS[doc_type]
        text1 = get_text_from_url(archived_url)
        text2 = get_text_from_url(current_url)
        comparison = compare_and_align_lines(text1, text2)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f'comparisons/fixed_diff_{doc_type}_{timestamp}.xlsx'
        os.makedirs('comparisons', exist_ok=True)
        save_to_excel(comparison, file_path)

        return send_file(file_path, as_attachment=True)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- GET / --- Default health route ---
@app.route('/')
def home():
    return "WebDiff API is running. Use /compare, /compare-fixed, or /get-doc-types"

# --- Run app ---
if __name__ == '__main__':
    app.run(debug=True)

